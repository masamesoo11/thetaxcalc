#!/usr/bin/env python3
"""
SEO Performance Analysis Dashboard Generator
For thetaxcalc.com - Google Search Console Data
Date Range: May 29-30, 2026 (24 hours)
Output: /home/z/my-project/upload/SEO-Analysis-thetaxcalc-2026-05-30.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_PATH = "/home/z/my-project/upload/SEO-Analysis-thetaxcalc-2026-05-30.xlsx"

# ============================================================
# COLOR PALETTE & STYLES
# ============================================================
DARK_BG = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_BG = PatternFill(start_color="2C3E6B", end_color="2C3E6B", fill_type="solid")
ACCENT_BG = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
GREEN_BG = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
RED_BG = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
YELLOW_BG = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
LIGHT_BG = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
LIGHT_GREEN_BG = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
LIGHT_RED_BG = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
LIGHT_YELLOW_BG = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
WHITE_BG = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GRAY_BG = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
WHITE_FONT_LG = Font(color="FFFFFF", bold=True, size=14)
WHITE_FONT_XL = Font(color="FFFFFF", bold=True, size=18)
WHITE_FONT_SM = Font(color="FFFFFF", size=10)
DARK_FONT = Font(color="1B2A4A", bold=True, size=11)
DARK_FONT_LG = Font(color="1B2A4A", bold=True, size=14)
DARK_FONT_SM = Font(color="1B2A4A", size=10)
RED_FONT = Font(color="DC2626", bold=True, size=11)
GREEN_FONT = Font(color="059669", bold=True, size=11)
YELLOW_FONT = Font(color="D97706", bold=True, size=11)
LINK_FONT = Font(color="3B82F6", size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RTL = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, data, fill=HEADER_BG, font=WHITE_FONT):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def write_data_row(ws, row, data, alt=False, fonts=None):
    fill = GRAY_BG if alt else WHITE_BG
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if fonts and col <= len(fonts) and fonts[col-1]:
            cell.font = fonts[col-1]
        else:
            cell.font = DARK_FONT_SM


def write_title(ws, row, col, text, merge_end_col=None, font=WHITE_FONT_XL, fill=DARK_BG):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font
    cell.fill = fill
    cell.alignment = CENTER
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = THIN_BORDER


def write_kpi_box(ws, row, col, label, value, status="neutral", merge_width=2):
    """Write a KPI box with label and value"""
    status_colors = {
        "good": (GREEN_BG, WHITE_FONT),
        "bad": (RED_BG, WHITE_FONT),
        "warning": (YELLOW_BG, WHITE_FONT),
        "neutral": (ACCENT_BG, WHITE_FONT),
        "info": (HEADER_BG, WHITE_FONT),
    }
    fill, font = status_colors.get(status, (ACCENT_BG, WHITE_FONT))
    
    # Label row
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + merge_width - 1)
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(color="FFFFFF", size=9)
    label_cell.fill = fill
    label_cell.alignment = CENTER
    label_cell.border = THIN_BORDER
    for c in range(col, col + merge_width):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = THIN_BORDER
    
    # Value row
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col + merge_width - 1)
    val_cell = ws.cell(row=row+1, column=col, value=value)
    val_cell.font = Font(color="FFFFFF", bold=True, size=16)
    val_cell.fill = fill
    val_cell.alignment = CENTER
    val_cell.border = THIN_BORDER
    for c in range(col, col + merge_width):
        ws.cell(row=row+1, column=c).fill = fill
        ws.cell(row=row+1, column=c).border = THIN_BORDER


# ============================================================
# DATA
# ============================================================

hourly_data = [
    ("14:00", 0, 5, 0, 34.2),
    ("15:00", 0, 7, 0, 71.4),
    ("16:00", 2, 7, 28.57, 32.0),
    ("17:00", 0, 5, 0, 30.0),
    ("18:00", 0, 5, 0, 21.6),
    ("19:00", 0, 13, 0, 30.6),
    ("20:00", 0, 5, 0, 51.2),
    ("21:00", 0, 7, 0, 34.4),
    ("22:00", 0, 11, 0, 38.0),
    ("23:00", 1, 12, 8.33, 27.2),
    ("00:00", 0, 8, 0, 9.6),
    ("01:00", 0, 7, 0, 44.4),
    ("02:00", 0, 22, 0, 32.7),
    ("03:00", 0, 11, 0, 18.8),
    ("04:00", 0, 20, 0, 35.6),
    ("05:00", 0, 17, 0, 24.6),
    ("06:00", 1, 21, 4.76, 36.8),
    ("07:00", 0, 43, 0, 40.3),
    ("08:00", 0, 21, 0, 37.9),
    ("09:00", 0, 9, 0, 19.7),
    ("10:00", 0, 30, 0, 34.0),
    ("11:00", 0, 9, 0, 51.1),
    ("12:00", 0, 1, 0, 11.0),
    ("13:00", 0, 0, 0, 0),
]

pages_data = [
    ("/new-york-tax-calculator", 2, 116, 1.72, 39.84),
    ("/", 1, 22, 4.55, 11.77),
    ("/compare/california-vs-new-york", 1, 8, 12.5, 17.0),
    ("/compare/texas-vs-california", 1, 5, 20.0, 23.6),
    ("/compare/texas-vs-new-york", 1, 2, 50.0, 30.0),
    ("/illinois-tax-calculator", 0, 46, 0, 41.57),
    ("/compare", 0, 28, 0, 28.07),
    ("/self-employment-tax-calculator", 0, 18, 0, 64.17),
    ("/salary/85000", 0, 11, 0, 13.82),
    ("/salary/55000", 0, 11, 0, 17.91),
    ("/salary/75000", 0, 8, 0, 13.5),
    ("/salary/60000", 0, 5, 0, 8.0),
    ("/salary/70000", 0, 5, 0, 19.6),
    ("/salary/100000", 0, 4, 0, 26.75),
    ("/compare/florida-vs-texas", 0, 3, 0, 9.0),
    ("/compare/florida-vs-california", 0, 3, 0, 17.67),
    ("/salary/90000", 0, 3, 0, 15.67),
    ("/glossary", 1, 2, 50.0, 3.5),
    ("/salary/40000", 0, 2, 0, 9.0),
    ("/salary/50000", 0, 2, 0, 15.0),
    ("/compare/florida-vs-new-york", 0, 2, 0, 21.5),
    ("/compare/texas-vs-illinois", 0, 2, 0, 78.5),
    ("/salary/80000", 0, 1, 0, 6.0),
    ("/salary/30000", 0, 1, 0, 9.0),
    ("/salary/120000", 0, 1, 0, 11.0),
    ("/salary/65000", 0, 1, 0, 12.0),
    ("/salary/45000", 0, 1, 0, 18.0),
    ("/privacy", 0, 1, 0, 2.0),
    ("/terms", 0, 1, 0, 5.0),
]

# Query data - categorized
queries_calculator = [
    ("tax calculator 2026 nyc", 2, 30.5, "حاسبة ضرائب NYC"),
    ("ny taxes calculator", 2, 52.5, "حاسبة ضرائب نيويورك"),
    ("new york tax income calculator", 2, 56.5, "حاسبة ضريبة دخل نيويورك"),
    ("tax in new york state calculator", 2, 80.5, "حاسبة ضريبة ولاية نيويورك"),
    ("nyc paycheck tax", 1, 53.0, "ضريبة راتب NYC"),
    ("nyc tax calculator", 1, 57.0, "حاسبة ضرائب NYC"),
    ("ny income tax calculator", 1, 58.0, "حاسبة ضريبة دخل NY"),
    ("new york income tax calculator", 1, 79.0, "حاسبة ضريبة دخل نيويورك"),
    ("new york income calculator", 1, 52.0, "حاسبة دخل نيويورك"),
    ("calculate ny state tax", 1, 44.0, "حساب ضريبة ولاية نيويورك"),
    ("calculate ny tax", 1, 51.0, "حساب ضريبة نيويورك"),
    ("income tax calculator illinois", 2, 42.5, "حاسبة ضريبة دخل إلينوي"),
    ("illinois tax calculator", 2, 49.0, "حاسبة ضرائب إلينوي"),
    ("illinois paycheck tax calculator", 2, 52.5, "حاسبة ضريبة راتب إلينوي"),
    ("adp tax calculator illinois", 1, 35.0, "حاسبة ADP ضرائب إلينوي"),
    ("calculate paycheck illinois", 1, 53.0, "حساب راتب إلينوي"),
    ("how much tax is taken out of paycheck in illinois", 1, 66.0, "كم ضريبة تخصم من راتب إلينوي"),
    ("tax calculator 2026", 1, 68.0, "حاسبة ضرائب 2026"),
    ("federal and state tax rate calculator", 1, 65.0, "حاسبة معدل الضرائب الفيدرالية والولائية"),
    ("tax calculator", 1, 82.0, "حاسبة ضرائب"),
]

queries_self_employment = [
    ("federal self employment tax calculator", 1, 63.0, "حاسبة ضريبة العمل الحر الفيدرالية"),
    ("self employment income tax calculator", 1, 65.0, "حاسبة ضريبة دخل العمل الحر"),
    ("tax self employed calculator", 1, 67.0, "حاسبة ضرائب المستقلين"),
    ("calculator self employment tax", 1, 70.0, "حاسبة ضريبة العمل الحر"),
    ("tax calculator with self employment", 1, 70.0, "حاسبة ضرائب مع العمل الحر"),
    ("self employment tax return calculator", 1, 72.0, "حاسبة إقرار ضريبة العمل الحر"),
    ("calculate se tax", 1, 75.0, "حساب ضريبة SE"),
    ("self employed tax return calculator", 1, 76.0, "حاسبة إقرار ضريبة المستقلين"),
    ("self employment tax calculation", 1, 87.0, "حساب ضريبة العمل الحر"),
    ("employment and self employment tax calculator", 1, 90.0, "حاسبة ضرائب التوظيف والعمل الحر"),
    ("calculator for self employment tax", 1, 91.0, "حاسبة لضريبة العمل الحر"),
    ("taxes on self employment calculator", 1, 95.0, "حاسبة ضرائب العمل الحر"),
    ("self employed federal tax rate", 1, 99.0, "معدل الضريبة الفيدرالية للمستقلين"),
]

queries_comparison = [
    ("california taxes vs new york", 3, 19.33, "ضرائب كاليفورنيا مقابل نيويورك"),
    ("louisiana vs. texas state taxes", 1, 43.0, "ضرائب لويزيانا مقابل تكساس"),
    ("illinois state tax vs federal tax", 1, 45.0, "ضريبة إلينوي مقابل الفيدرالية"),
    ("illinois taxes compared to other states", 1, 88.0, "ضرائب إلينوي مقارنة بولايات أخرى"),
    ("compare illinois state tax credits", 1, 24.0, "مقارنة أرصدة ضريبة إلينوي"),
]

queries_general = [
    ("smartasset nyc paycheck calculator", 1, 49.0, "حاسبة راتب smartasset NYC"),
    ("how to find the tax", 1, 93.0, "كيفية حساب الضريبة"),
    ("how much i will earn after tax", 1, 95.0, "كم سأكسب بعد الضريبة"),
]

countries_data = [
    ("ليبيا 🇱🇾 (مستخدم الموقع)", 3, 3, 100.0, 1.0),
    ("الولايات المتحدة 🇺🇸 (السوق الرئيسي)", 1, 268, 0.37, 36.67),
    ("المملكة المتحدة 🇬🇧", 0, 3, 0, 31.33),
    ("الأرجنتين 🇦🇷", 0, 2, 0, 4.5),
    ("المكسيك 🇲🇽", 0, 2, 0, 5.5),
]

devices_data = [
    ("سطح المكتب (Desktop)", 3, 247, 1.21, 38.91),
    ("الهاتف الجوال (Mobile)", 1, 49, 2.04, 12.14),
    ("جهاز لوحي (Tablet)", 0, 9, 0, 4.67),
]

# Opportunities / Recommendations
opportunities = [
    (
        "حرج",
        "تحسين عنوان ووصف صفحة حاسبة إلينوي",
        "/illinois-tax-calculator",
        "46 ظهور و0 نقرات! الموضع 41.57 - العنوان أو الوصف لا يجذب المستخدمين. يجب إعادة كتابة title tag وmeta description بشكل عاجل.",
        "متوقع زيادة CTR إلى 2-5% = 1-2 نقرات إضافية يومياً",
        "فوري"
    ),
    (
        "حرج",
        "تحسين صفحة الرواتب /salary/85000",
        "/salary/85000",
        "الموضع 13.82 (قريب من الصفحة الأولى!) لكن 0 نقرات. المشكلة في meta description - لا يبدو جذاباً في نتائج البحث.",
        "الوصول للصفحة الأولى يمكن أن يضاعف الظهور والنقرات",
        "خلال 3 أيام"
    ),
    (
        "حرج",
        "تحسين ترتيب صفحة حاسبة نيويورك",
        "/new-york-tax-calculator",
        "أعلى صفحة ظهوراً (116 ظهور) لكن الموضع 39.84 يعني ظهور في الصفحة 4! يجب تحسين المحتوى وبناء روابط.",
        "الانتقال من الموضع 40 إلى 20 يمكن أن يزيد النقرات 10 أضعاف",
        "خلال أسبوعين"
    ),
    (
        "عالي",
        "الاستثمار في صفحات المقارنة",
        "/compare/*",
        "أعلى CTR بين جميع الصفحات (12.5%-50%)! هذا تنسيق ناجح. يجب إنشاء المزيد من صفحات المقارنة.",
        "كل صفحة مقارنة جديدة = 5-15 نقرات إضافية محتملة",
        "خلال أسبوع"
    ),
    (
        "عالي",
        "تحسين صفحة حاسبة العمل الحر",
        "/self-employment-tax-calculator",
        "18 ظهور لكن الموضع 64.17 (صفحة 7!) - تحتاج تحسين SEO شامل: محتوى أطول، كلمات مفتاحية، روابط داخلية.",
        "الوصول للموضع 20-30 يمكن أن يجلب 2-5 نقرات يومياً",
        "خلال أسبوعين"
    ),
    (
        "عالي",
        "تحسين تجربة الموبايل أولاً",
        "الموقع بالكامل",
        "Mobile CTR 2.04% vs Desktop 1.21% وموضع أفضل (12.14 vs 38.91). يجب إعطاء الأولوية لتحسين الموبايل.",
        "تحسين Core Web Vitals للموبايل يرفع الترتيب",
        "مستمر"
    ),
    (
        "متوسط",
        "إنشاء محتوى لكلمات العمل الحر",
        "محتوى جديد",
        "13 استعلام عن self-employment tax لكن الموضع 63-99. يجب إنشاء مقالات دليلية شاملة حول الموضوع.",
        "استهداف الكلمات الطويلة (long-tail) بمراتب أقل",
        "خلال شهر"
    ),
    (
        "متوسط",
        "تحسين الصفحة الرئيسية",
        "/",
        "22 ظهور وموضع 11.77 - قريبة من الصفحة الأولى! تحسين بسيط يمكن أن يدخلها أعلى 10.",
        "دخول أعلى 10 = زيادة كبيرة في النقرات",
        "خلال أسبوع"
    ),
    (
        "متوسط",
        "استهداف الكلمات المفتاحية ذات الحجم الأعلى",
        "استراتيجية المحتوى",
        "معظم الاستعلامات impressions منخفضة (1-3). يجب استهداف كلمات بحجم بحث أعلى مثل 'tax calculator', 'paycheck calculator'.",
        "الوصول لكلمات أعلى حجم = زيادة الظهور 5-10 أضعاف",
        "خلال شهرين"
    ),
    (
        "منخفض",
        "إضافة Schema Markup",
        "الموقع بالكامل",
        "إضافة FinancialProduct أو SoftwareApplication schema لتحسين ظهور النتائج مع rich snippets.",
        "Rich snippets تزيد CTR بنسبة 20-30%",
        "خلال أسبوع"
    ),
    (
        "منخفض",
        "تحسين سرعة الموقع",
        "الموقع بالكامل",
        "Core Web Vitals مهمة خاصة للموبايل. يجب قياس وتحسين LCP, FID, CLS.",
        "تحسين السرعة يرفع الترتيب بشكل عام",
        "مستمر"
    ),
    (
        "منخفض",
        "بناء روابط خلفية (Backlinks)",
        "الموقع بالكامل",
        "الموقع جديد ويحتاج بناء سلطة. التركيز على مواقع الضرائب والمالية.",
        "كل backlink عالي الجودة يزيد Domain Authority",
        "خلال 3 أشهر"
    ),
]


# ============================================================
# CREATE WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: لوحة المعلومات (Summary Dashboard)
# ============================================================
ws1 = wb.active
ws1.title = "لوحة المعلومات"
ws1.sheet_properties.tabColor = "1B2A4A"
set_col_widths(ws1, [3, 22, 18, 18, 18, 18, 18, 18, 18, 18, 3])

# Title banner
for c in range(1, 11):
    ws1.cell(row=1, column=c).fill = DARK_BG
    ws1.cell(row=1, column=c).border = THIN_BORDER

ws1.merge_cells('B2:I2')
cell = ws1.cell(row=2, column=2, value="📊 تحليل أداء SEO - thetaxcalc.com")
cell.font = Font(color="FFFFFF", bold=True, size=22)
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 11):
    ws1.cell(row=2, column=c).fill = DARK_BG
    ws1.cell(row=2, column=c).border = THIN_BORDER

ws1.merge_cells('B3:I3')
cell = ws1.cell(row=3, column=2, value="فترة البيانات: 29-30 مايو 2026 (24 ساعة) | مصدر البيانات: Google Search Console")
cell.font = Font(color="94A3B8", size=11)
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 11):
    ws1.cell(row=3, column=c).fill = DARK_BG
    ws1.cell(row=3, column=c).border = THIN_BORDER

# Row 4: Spacer
for c in range(1, 11):
    ws1.cell(row=4, column=c).fill = WHITE_BG

# KPI Row 1 (row 5-6)
write_kpi_box(ws1, 5, 2, "إجمالي النقرات", "4", "bad", 2)
write_kpi_box(ws1, 5, 4, "إجمالي الظهور", "305", "neutral", 2)
write_kpi_box(ws1, 5, 6, "متوسط نسبة النقر CTR", "1.31%", "bad", 2)
write_kpi_box(ws1, 5, 8, "متوسط الموضع", "36.84", "bad", 2)

# KPI Row 2 (row 7-8)
write_kpi_box(ws1, 7, 2, "نقرات السوق الرئيسي (USA)", "1", "bad", 2)
write_kpi_box(ws1, 7, 4, "CTR السوق الأمريكي", "0.37%", "bad", 2)
write_kpi_box(ws1, 7, 6, "أفضل صفحة ظهوراً", "NY Calculator", "warning", 2)
write_kpi_box(ws1, 7, 8, "أعلى CTR", "صفحات المقارنة", "good", 2)

# Spacer
for c in range(1, 11):
    ws1.cell(row=9, column=c).fill = WHITE_BG

# Assessment Section
ws1.merge_cells('B10:I10')
cell = ws1.cell(row=10, column=2, value="⚠️ التقييم العام: الموقع في مراحله الأولى - يحتاج تحسينات عاجلة")
cell.font = Font(color="1B2A4A", bold=True, size=14)
cell.fill = LIGHT_YELLOW_BG
cell.alignment = CENTER
cell.border = THIN_BORDER
for c in range(2, 10):
    ws1.cell(row=10, column=c).fill = LIGHT_YELLOW_BG
    ws1.cell(row=10, column=c).border = THIN_BORDER

# Key Findings
ws1.merge_cells('B12:I12')
cell = ws1.cell(row=12, column=2, value="🔍 النتائج الرئيسية")
cell.font = DARK_FONT_LG
cell.fill = LIGHT_BG
cell.alignment = CENTER
cell.border = THIN_BORDER
for c in range(2, 10):
    ws1.cell(row=12, column=c).fill = LIGHT_BG
    ws1.cell(row=12, column=c).border = THIN_BORDER

findings = [
    ("🔴", "أزمة CTR في السوق الأمريكي", "CTR 0.37% فقط في الولايات المتحدة - أقل بكثير من المتوسط الصناعي (3-5%). معظم الصفحات تظهر بعد الموضع 30."),
    ("🔴", "صفحة إلينوي بدون نقرات", "46 ظهور و0 نقرات! هذا يعني أن العنوان أو الوصف غير جذابين على الإطلاق. تحتاج إعادة كتابة عاجلة."),
    ("🟡", "صفحات الرواتب قريبة من الصفحة الأولى", "/salary/85000 في الموضع 13.82 و /salary/60000 في الموضع 8.0 - تحسين بسيط يمكن أن يدخلها أعلى 10."),
    ("🟢", "صفحات المقارنة هي النجم", "أعلى CTR بين جميع الصفحات (12.5%-50%). هذا التنسيق يحبه المستخدمون ويجب التوسع فيه."),
    ("🟢", "الموبايل يتفوق على الديسكتوب", "Mobile CTR 2.04% vs Desktop 1.21% وموضع أفضل (12.14 vs 38.91). يجب إعطاء الأولوية للتجربة المحمولة."),
    ("🔴", "حاسبة العمل الحر في مكان سيء", "18 ظهور لكن الموضع 64.17 - في الصفحة السابعة! تحتاج تحسين SEO شامل."),
]

for i, (emoji, title, desc) in enumerate(findings):
    row = 14 + i * 2
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    cell = ws1.cell(row=row, column=2, value=f"{emoji} {title}")
    cell.font = Font(color="1B2A4A", bold=True, size=12)
    cell.fill = WHITE_BG
    cell.alignment = RTL
    cell.border = THIN_BORDER
    for c in range(2, 10):
        ws1.cell(row=row, column=c).fill = WHITE_BG
        ws1.cell(row=row, column=c).border = THIN_BORDER
    
    ws1.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=9)
    cell = ws1.cell(row=row+1, column=2, value=desc)
    cell.font = Font(color="475569", size=10)
    cell.fill = LIGHT_BG
    cell.alignment = RTL
    cell.border = THIN_BORDER
    for c in range(2, 10):
        ws1.cell(row=row+1, column=c).fill = LIGHT_BG
        ws1.cell(row=row+1, column=c).border = THIN_BORDER

# Device comparison section
dev_start = 28
ws1.merge_cells(f'B{dev_start}:I{dev_start}')
cell = ws1.cell(row=dev_start, column=2, value="📱 مقارنة الأجهزة")
cell.font = DARK_FONT_LG
cell.fill = LIGHT_BG
cell.alignment = CENTER
cell.border = THIN_BORDER
for c in range(2, 10):
    ws1.cell(row=dev_start, column=c).fill = LIGHT_BG
    ws1.cell(row=dev_start, column=c).border = THIN_BORDER

dev_headers = ["الجهاز", "النقرات", "الظهور", "CTR", "الموضع", "التقييم"]
write_header_row(ws1, dev_start + 1, dev_headers, ACCENT_BG)
# Adjust to use columns B-G
for i, (device, clicks, imp, ctr, pos) in enumerate(devices_data):
    row = dev_start + 2 + i
    data = [device, clicks, imp, f"{ctr}%", f"{pos:.2f}", ""]
    alt = i % 2 == 1
    fill = GRAY_BG if alt else WHITE_BG
    
    assessment = ""
    if "Mobile" in device:
        assessment = "⭐ الأفضل"
    elif "Desktop" in device:
        assessment = "⚠️ يحتاج تحسين"
    else:
        assessment = "➖ محدود"
    data[5] = assessment
    
    for col, val in enumerate(data, 2):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.font = DARK_FONT_SM

# Country section
country_start = dev_start + 6
ws1.merge_cells(f'B{country_start}:I{country_start}')
cell = ws1.cell(row=country_start, column=2, value="🌍 توزيع الدول")
cell.font = DARK_FONT_LG
cell.fill = LIGHT_BG
cell.alignment = CENTER
cell.border = THIN_BORDER
for c in range(2, 10):
    ws1.cell(row=country_start, column=c).fill = LIGHT_BG
    ws1.cell(row=country_start, column=c).border = THIN_BORDER

country_headers = ["الدولة", "النقرات", "الظهور", "CTR", "الموضع", "ملاحظة"]
write_header_row(ws1, country_start + 1, country_headers, ACCENT_BG)
for i, (country, clicks, imp, ctr, pos) in enumerate(countries_data):
    row = country_start + 2 + i
    note = ""
    if "ليبيا" in country:
        note = "⛔ عمليات المالك - لا تحسب!"
    elif "الولايات" in country:
        note = "🎯 السوق المستهدف - CTR كارثي"
    else:
        note = "غير مستهدف"
    
    data = [country, clicks, imp, f"{ctr}%", f"{pos:.2f}", note]
    alt = i % 2 == 1
    fill = GRAY_BG if alt else WHITE_BG
    for col, val in enumerate(data, 2):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.font = DARK_FONT_SM

# ============================================================
# SHEET 2: أداء الصفحات (Page Performance)
# ============================================================
ws2 = wb.create_sheet("أداء الصفحات")
ws2.sheet_properties.tabColor = "3B82F6"
set_col_widths(ws2, [3, 40, 12, 14, 12, 12, 14, 35, 3])

# Title
ws2.merge_cells('B1:H1')
cell = ws2.cell(row=1, column=2, value="📄 أداء الصفحات - thetaxcalc.com")
cell.font = WHITE_FONT_XL
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 9):
    ws2.cell(row=1, column=c).fill = DARK_BG

ws2.merge_cells('B2:H2')
cell = ws2.cell(row=2, column=2, value="مرتبة حسب عدد الظهور | 29-30 مايو 2026")
cell.font = WHITE_FONT_SM
cell.fill = HEADER_BG
cell.alignment = CENTER
for c in range(1, 9):
    ws2.cell(row=2, column=c).fill = HEADER_BG

# Headers
page_headers = [
    "الصفحة", "النقرات", "الظهور", "CTR", "الموضع",
    "الأولوية", "التقييم والتوصية"
]
write_header_row(ws2, 4, page_headers, ACCENT_BG)

for i, (page, clicks, imp, ctr, pos) in enumerate(pages_data):
    row = 5 + i
    alt = i % 2 == 1
    
    # Determine priority and assessment
    if page == "/illinois-tax-calculator":
        priority = "🔴 حرج"
        assessment = "46 ظهور بدون نقرات! إعادة كتابة العنوان والوصف فوراً"
    elif page == "/new-york-tax-calculator":
        priority = "🔴 حرج"
        assessment = "أعلى ظهور لكن الموضع 40+ = صفحة 4-5. يحتاج بناء روابط وتحسين محتوى"
    elif page == "/self-employment-tax-calculator":
        priority = "🟠 عالي"
        assessment = "الموضع 64 = صفحة 7. يحتاج محتوى شامل وكلمات مفتاحية"
    elif page == "/salary/85000":
        priority = "🟠 عالي"
        assessment = "موضع 13.82 قريب من الصفحة الأولى! تحسين meta description فقط"
    elif page == "/salary/60000":
        priority = "🟡 متوسط"
        assessment = "موضع 8.0 ممتاز! لكن impressions منخفضة. زيادة المحتوى"
    elif "/compare/" in page:
        priority = "🟢 استمرار"
        assessment = f"CTR {ctr}% ممتاز! تنسيق ناجح - إنشاء المزيد"
    elif page == "/":
        priority = "🟡 متوسط"
        assessment = "موضع 11.77 قريب من أعلى 10. تحسين بسيط يمكن أن يحدث فرقاً"
    elif "/salary/" in page:
        priority = "🟡 متوسط"
        assessment = f"موضع {pos:.1f} - إضافة محتوى ثري وكلمات مفتاحية"
    elif page == "/compare":
        priority = "🟡 متوسط"
        assessment = "28 ظهور بدون نقرات - تحسين العنوان والربط الداخلي"
    elif page == "/glossary":
        priority = "🟢 جيد"
        assessment = "CTR 50% وموضع 3.5! لكن ظهور منخفض - إضافة مصطلحات أكثر"
    else:
        priority = "⚪ منخفض"
        assessment = "أولوية منخفضة حالياً"
    
    data = [page, clicks, imp, f"{ctr}%", f"{pos:.2f}", priority, assessment]
    fill = GRAY_BG if alt else WHITE_BG
    
    for col, val in enumerate(data, 2):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.font = DARK_FONT_SM
        cell.alignment = CENTER if col < 7 else LEFT
        if col == 2:
            cell.font = LINK_FONT
            cell.alignment = LEFT
        if col == 7 and "حرج" in str(val):
            cell.font = RED_FONT
        elif col == 7 and "عالي" in str(val):
            cell.font = YELLOW_FONT
        elif col == 7 and "استمرار" in str(val):
            cell.font = GREEN_FONT

# Summary stats at bottom
summary_row = 5 + len(pages_data) + 2
ws2.merge_cells(f'B{summary_row}:H{summary_row}')
cell = ws2.cell(row=summary_row, column=2, value="📊 ملخص أداء الصفحات")
cell.font = WHITE_FONT
cell.fill = HEADER_BG
cell.alignment = CENTER
for c in range(2, 9):
    ws2.cell(row=summary_row, column=c).fill = HEADER_BG

total_clicks = sum(p[1] for p in pages_data)
total_imp = sum(p[2] for p in pages_data)
pages_with_clicks = sum(1 for p in pages_data if p[1] > 0)
pages_zero_clicks = sum(1 for p in pages_data if p[1] == 0 and p[2] > 0)
pages_top10 = sum(1 for p in pages_data if p[4] <= 10)
pages_beyond_30 = sum(1 for p in pages_data if p[4] > 30 and p[2] > 0)

stats = [
    ("إجمالي الصفحات", len(pages_data)),
    ("صفحات بنقرات", f"{pages_with_clicks} ({pages_with_clicks/len(pages_data)*100:.0f}%)"),
    ("صفحات بدون نقرات", f"{pages_zero_clicks} ({pages_zero_clicks/len(pages_data)*100:.0f}%)"),
    ("صفحات في أعلى 10", f"{pages_top10} صفحة"),
    ("صفحات بعد الموضع 30", f"{pages_beyond_30} صفحة (خارج الصفحات الأولى)"),
    ("إجمالي النقرات", total_clicks),
    ("إجمالي الظهور", total_imp),
    ("متوسط CTR", f"{total_clicks/total_imp*100:.2f}%"),
]

for i, (label, val) in enumerate(stats):
    r = summary_row + 1 + i
    ws2.merge_cells(f'B{r}:D{r}')
    cell = ws2.cell(row=r, column=2, value=label)
    cell.font = DARK_FONT
    cell.fill = LIGHT_BG
    cell.alignment = RTL
    cell.border = THIN_BORDER
    for c in range(2, 5):
        ws2.cell(row=r, column=c).fill = LIGHT_BG
        ws2.cell(row=r, column=c).border = THIN_BORDER
    
    ws2.merge_cells(f'E{r}:H{r}')
    cell = ws2.cell(row=r, column=5, value=str(val))
    cell.font = Font(color="1B2A4A", bold=True, size=12)
    cell.fill = WHITE_BG
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    for c in range(5, 9):
        ws2.cell(row=r, column=c).fill = WHITE_BG
        ws2.cell(row=r, column=c).border = THIN_BORDER


# ============================================================
# SHEET 3: تحليل الاستعلامات (Query Analysis)
# ============================================================
ws3 = wb.create_sheet("تحليل الاستعلامات")
ws3.sheet_properties.tabColor = "059669"
set_col_widths(ws3, [3, 35, 12, 12, 12, 30, 30, 3])

# Title
ws3.merge_cells('B1:G1')
cell = ws3.cell(row=1, column=2, value="🔎 تحليل استعلامات البحث - thetaxcalc.com")
cell.font = WHITE_FONT_XL
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 8):
    ws3.cell(row=1, column=c).fill = DARK_BG

ws3.merge_cells('B2:G2')
cell = ws3.cell(row=2, column=2, value="104 استعلام | تصنيف حسب القصد والوضوع | 0 نقرات تقريباً من جميع الاستعلامات")
cell.font = Font(color="EF4444", bold=True, size=11)
cell.fill = RED_BG
cell.alignment = CENTER
for c in range(1, 8):
    ws3.cell(row=2, column=c).fill = RED_BG

# Category: Tax Calculators
current_row = 4
categories = [
    ("🧮 حاسبات الضرائب (Tax Calculators)", queries_calculator, ACCENT_BG),
    ("💼 ضريبة العمل الحر (Self-Employment Tax)", queries_self_employment, YELLOW_BG),
    ("⚖️ مقارنات الضرائب (Tax Comparisons)", queries_comparison, GREEN_BG),
    ("📝 استعلامات عامة (General)", queries_general, HEADER_BG),
]

for cat_name, queries, cat_fill in categories:
    ws3.merge_cells(f'B{current_row}:G{current_row}')
    cell = ws3.cell(row=current_row, column=2, value=cat_name)
    cell.font = WHITE_FONT
    cell.fill = cat_fill
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    for c in range(2, 8):
        ws3.cell(row=current_row, column=c).fill = cat_fill
        ws3.cell(row=current_row, column=c).border = THIN_BORDER
    
    current_row += 1
    
    q_headers = ["الاستعلام (إنجليزي)", "الظهور", "الموضع", "الترجمة", "التقييم"]
    write_header_row(ws3, current_row, q_headers, PatternFill(start_color="475569", end_color="475569", fill_type="solid"))
    current_row += 1
    
    for j, (query, imp, pos, ar_translation) in enumerate(queries):
        alt = j % 2 == 1
        fill = GRAY_BG if alt else WHITE_BG
        
        if pos <= 20:
            assessment = "🟢 جيد - قريب من الصفحة الأولى"
        elif pos <= 30:
            assessment = "🟡 متوسط - يحتاج تحسين بسيط"
        elif pos <= 50:
            assessment = "🟠 ضعيف - يحتاج تحسين كبير"
        else:
            assessment = "🔴 سيء - يحتاج إعادة بناء المحتوى"
        
        data = [query, imp, f"{pos:.1f}", ar_translation, assessment]
        for col, val in enumerate(data, 2):
            cell = ws3.cell(row=current_row, column=col, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.font = DARK_FONT_SM
            cell.alignment = CENTER if col in [3, 4] else LEFT
            if col == 6 and "سيء" in str(val):
                cell.font = RED_FONT
            elif col == 6 and "ضعيف" in str(val):
                cell.font = YELLOW_FONT
            elif col == 6 and "جيد" in str(val):
                cell.font = GREEN_FONT
        
        current_row += 1
    
    current_row += 1  # Spacer between categories

# Query summary
current_row += 1
ws3.merge_cells(f'B{current_row}:G{current_row}')
cell = ws3.cell(row=current_row, column=2, value="📊 ملخص تحليل الاستعلامات")
cell.font = WHITE_FONT
cell.fill = HEADER_BG
cell.alignment = CENTER
for c in range(2, 8):
    ws3.cell(row=current_row, column=c).fill = HEADER_BG

current_row += 1
q_summary = [
    ("إجمالي الاستعلامات", "104 استعلام"),
    ("استعلامات بنقرات", "~0 (نقرات من بيانات مجمعة فقط)"),
    ("استعلامات في أعلى 20 موضع", f"{sum(1 for q in queries_calculator + queries_self_employment + queries_comparison + queries_general if q[2] <= 20)} استعلام"),
    ("استعلامات في الموضع 50+", f"{sum(1 for q in queries_calculator + queries_self_employment + queries_comparison + queries_general if q[2] > 50)} استعلام (الأغلبية!)"),
    ("أعلى حجم ظهور", "california taxes vs new york (3 impressions)"),
    ("نوع الاستعلامات", "طويلة الذيل (long-tail) بحجم بحث منخفض جداً"),
    ("الفرصة", "استهداف كلمات أقصر وأعلى حجم مثل 'tax calculator', 'paycheck calculator'"),
]

for label, val in q_summary:
    ws3.merge_cells(f'B{current_row}:D{current_row}')
    cell = ws3.cell(row=current_row, column=2, value=label)
    cell.font = DARK_FONT
    cell.fill = LIGHT_BG
    cell.alignment = RTL
    cell.border = THIN_BORDER
    for c in range(2, 5):
        ws3.cell(row=current_row, column=c).fill = LIGHT_BG
        ws3.cell(row=current_row, column=c).border = THIN_BORDER
    
    ws3.merge_cells(f'E{current_row}:G{current_row}')
    cell = ws3.cell(row=current_row, column=5, value=val)
    cell.font = Font(color="1B2A4A", size=10)
    cell.fill = WHITE_BG
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    for c in range(5, 8):
        ws3.cell(row=current_row, column=c).fill = WHITE_BG
        ws3.cell(row=current_row, column=c).border = THIN_BORDER
    
    current_row += 1


# ============================================================
# SHEET 4: الفرص (Opportunities)
# ============================================================
ws4 = wb.create_sheet("الفرص والتوصيات")
ws4.sheet_properties.tabColor = "D97706"
set_col_widths(ws4, [3, 14, 35, 28, 50, 35, 16, 3])

# Title
ws4.merge_cells('B1:G1')
cell = ws4.cell(row=1, column=2, value="🚀 الفرص والتوصيات SEO - thetaxcalc.com")
cell.font = WHITE_FONT_XL
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 8):
    ws4.cell(row=1, column=c).fill = DARK_BG

ws4.merge_cells('B2:G2')
cell = ws4.cell(row=2, column=2, value="مرتبة حسب الأولوية | 12 توصية محددة وقابلة للتنفيذ")
cell.font = WHITE_FONT_SM
cell.fill = HEADER_BG
cell.alignment = CENTER
for c in range(1, 8):
    ws4.cell(row=2, column=c).fill = HEADER_BG

# Headers
opp_headers = ["الأولوية", "التوصية", "الصفحة", "التفاصيل", "التأثير المتوقع", "الجدول الزمني"]
write_header_row(ws4, 4, opp_headers, ACCENT_BG)

for i, (priority, title, page, details, impact, timeline) in enumerate(opportunities):
    row = 5 + i
    alt = i % 2 == 1
    fill = GRAY_BG if alt else WHITE_BG
    
    data = [priority, title, page, details, impact, timeline]
    for col, val in enumerate(data, 2):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.font = DARK_FONT_SM
        cell.alignment = CENTER if col in [2, 7] else LEFT
        
        if col == 2:
            if "حرج" in str(val):
                cell.font = Font(color="DC2626", bold=True, size=11)
                cell.fill = LIGHT_RED_BG
            elif "عالي" in str(val):
                cell.font = Font(color="D97706", bold=True, size=11)
                cell.fill = LIGHT_YELLOW_BG
            elif "متوسط" in str(val):
                cell.font = Font(color="3B82F6", bold=True, size=11)
            else:
                cell.font = Font(color="6B7280", size=10)

# Priority summary
ps_row = 5 + len(opportunities) + 2
ws4.merge_cells(f'B{ps_row}:G{ps_row}')
cell = ws4.cell(row=ps_row, column=2, value="📊 ملخص الأولويات")
cell.font = WHITE_FONT
cell.fill = HEADER_BG
cell.alignment = CENTER
for c in range(2, 8):
    ws4.cell(row=ps_row, column=c).fill = HEADER_BG

urgent = sum(1 for o in opportunities if o[0] == "حرج")
high = sum(1 for o in opportunities if o[0] == "عالي")
medium = sum(1 for o in opportunities if o[0] == "متوسط")
low = sum(1 for o in opportunities if o[0] == "منخفض")

priority_summary = [
    ("🔴 حرج (فوري)", urgent, RED_BG, "يجب البدء بها خلال 24-48 ساعة"),
    ("🟠 عالي", high, YELLOW_BG, "يجب البدء بها خلال أسبوع"),
    ("🟡 متوسط", medium, ACCENT_BG, "يجب البدء بها خلال 2-4 أسابيع"),
    ("⚪ منخفض", low, PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid"), "تحسينات مستمرة على المدى الطويل"),
]

for i, (label, count, pfill, desc) in enumerate(priority_summary):
    r = ps_row + 1 + i
    ws4.merge_cells(f'B{r}:C{r}')
    cell = ws4.cell(row=r, column=2, value=f"{label}: {count}")
    cell.font = WHITE_FONT
    cell.fill = pfill
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    for c in range(2, 4):
        ws4.cell(row=r, column=c).fill = pfill
        ws4.cell(row=r, column=c).border = THIN_BORDER
    
    ws4.merge_cells(f'D{r}:G{r}')
    cell = ws4.cell(row=r, column=4, value=desc)
    cell.font = DARK_FONT_SM
    cell.fill = WHITE_BG
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    for c in range(4, 8):
        ws4.cell(row=r, column=c).fill = WHITE_BG
        ws4.cell(row=r, column=c).border = THIN_BORDER


# ============================================================
# SHEET 5: الرسوم البيانية (Charts)
# ============================================================
ws5 = wb.create_sheet("الرسوم البيانية")
ws5.sheet_properties.tabColor = "8B5CF6"
set_col_widths(ws5, [3, 30, 12, 12, 12, 12, 3])

# Title
ws5.merge_cells('B1:F1')
cell = ws5.cell(row=1, column=2, value="📊 الرسوم البيانية - thetaxcalc.com")
cell.font = WHITE_FONT_XL
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 7):
    ws5.cell(row=1, column=c).fill = DARK_BG

# ---- Data for Charts ----

# Top 10 pages by impressions
chart_data_start = 3
write_header_row(ws5, chart_data_start, ["الصفحة", "الظهور", "النقرات", "CTR%", "الموضع"], ACCENT_BG)
top_pages = sorted(pages_data, key=lambda x: x[2], reverse=True)[:10]
for i, (page, clicks, imp, ctr, pos) in enumerate(top_pages):
    row = chart_data_start + 1 + i
    data = [page.replace("/", "").replace("-", " ")[:25], imp, clicks, ctr, pos]
    write_data_row(ws5, row, data, i % 2 == 1)

# Bar Chart: Top Pages by Impressions
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "أعلى 10 صفحات من حيث الظهور (Impressions)"
chart1.y_axis.title = "عدد الظهور"
chart1.x_axis.title = "الصفحة"
chart1.width = 28
chart1.height = 16

data_ref = Reference(ws5, min_col=3, min_row=chart_data_start, max_row=chart_data_start + 10)
cats_ref = Reference(ws5, min_col=2, min_row=chart_data_start + 1, max_row=chart_data_start + 10)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4

# Color the bars
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
s = chart1.series[0]
s.graphicalProperties.solidFill = "3B82F6"

ws5.add_chart(chart1, "H3")

# ---- Device Clicks Data for Pie Chart ----
pie_data_start = chart_data_start + 12
write_header_row(ws5, pie_data_start, ["الجهاز", "النقرات"], ACCENT_BG)
for i, (device, clicks, imp, ctr, pos) in enumerate(devices_data):
    row = pie_data_start + 1 + i
    write_data_row(ws5, row, [device, clicks], i % 2 == 1)

# Pie Chart: Clicks by Device
chart2 = PieChart()
chart2.title = "توزيع النقرات حسب الجهاز"
chart2.width = 18
chart2.height = 14

data_ref2 = Reference(ws5, min_col=3, min_row=pie_data_start, max_row=pie_data_start + 3)
cats_ref2 = Reference(ws5, min_col=2, min_row=pie_data_start + 1, max_row=pie_data_start + 3)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)

# Add data labels
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
chart2.dataLabels.showVal = True
chart2.dataLabels.showCatName = True

# Color slices
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties
pie_colors = ["3B82F6", "059669", "D97706"]
for idx, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart2.series[0].data_points.append(pt)

ws5.add_chart(chart2, "H20")

# ---- Hourly Performance Line Chart ----
hourly_data_start = pie_data_start + 5
write_header_row(ws5, hourly_data_start, ["الساعة", "الظهور", "النقرات"], ACCENT_BG)
for i, (hour, clicks, imp, ctr, pos) in enumerate(hourly_data):
    row = hourly_data_start + 1 + i
    write_data_row(ws5, row, [hour, imp, clicks], i % 2 == 1)

chart3 = LineChart()
chart3.title = "الأداء بالساعة (الظهور والنقرات)"
chart3.y_axis.title = "العدد"
chart3.x_axis.title = "الساعة (UTC)"
chart3.width = 28
chart3.height = 14
chart3.style = 10

data_ref3 = Reference(ws5, min_col=3, min_row=hourly_data_start, max_col=4, max_row=hourly_data_start + 24)
cats_ref3 = Reference(ws5, min_col=2, min_row=hourly_data_start + 1, max_row=hourly_data_start + 24)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)

# Style lines
chart3.series[0].graphicalProperties.line.solidFill = "3B82F6"
chart3.series[0].graphicalProperties.line.width = 25000
chart3.series[1].graphicalProperties.line.solidFill = "EF4444"
chart3.series[1].graphicalProperties.line.width = 25000

ws5.add_chart(chart3, "H38")

# ---- CTR by Page Type Bar Chart ----
ctr_data_start = hourly_data_start + 26
# Group pages by type
page_types = {
    "حاسبات الضرائب": [],
    "صفحات المقارنة": [],
    "صفحات الرواتب": [],
    "أخرى": [],
}
for page, clicks, imp, ctr, pos in pages_data:
    if "/compare/" in page:
        page_types["صفحات المقارنة"].append((clicks, imp, ctr))
    elif "/salary/" in page:
        page_types["صفحات الرواتب"].append((clicks, imp, ctr))
    elif "calculator" in page:
        page_types["حاسبات الضرائب"].append((clicks, imp, ctr))
    else:
        page_types["أخرى"].append((clicks, imp, ctr))

write_header_row(ws5, ctr_data_start, ["نوع الصفحة", "متوسط CTR%"], ACCENT_BG)
for i, (ptype, pages) in enumerate(page_types.items()):
    if pages:
        avg_ctr = sum(p[2] for p in pages) / len(pages)
    else:
        avg_ctr = 0
    row = ctr_data_start + 1 + i
    write_data_row(ws5, row, [ptype, round(avg_ctr, 2)], i % 2 == 1)

chart4 = BarChart()
chart4.type = "col"
chart4.style = 10
chart4.title = "متوسط CTR حسب نوع الصفحة"
chart4.y_axis.title = "متوسط CTR %"
chart4.width = 18
chart4.height = 14

data_ref4 = Reference(ws5, min_col=3, min_row=ctr_data_start, max_row=ctr_data_start + 4)
cats_ref4 = Reference(ws5, min_col=2, min_row=ctr_data_start + 1, max_row=ctr_data_start + 4)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
s4 = chart4.series[0]
s4.graphicalProperties.solidFill = "059669"

ws5.add_chart(chart4, "H55")

# ============================================================
# SHEET 6: الأداء بالساعة (Hourly Performance)
# ============================================================
ws6 = wb.create_sheet("الأداء بالساعة")
ws6.sheet_properties.tabColor = "EF4444"
set_col_widths(ws6, [3, 12, 12, 14, 12, 12, 30, 3])

# Title
ws6.merge_cells('B1:G1')
cell = ws6.cell(row=1, column=2, value="⏰ الأداء بالساعة - thetaxcalc.com")
cell.font = WHITE_FONT_XL
cell.fill = DARK_BG
cell.alignment = CENTER
for c in range(1, 8):
    ws6.cell(row=1, column=c).fill = DARK_BG

# Headers
hourly_headers = ["الساعة", "النقرات", "الظهور", "CTR", "الموضع", "ملاحظة"]
write_header_row(ws6, 3, hourly_headers, ACCENT_BG)

for i, (hour, clicks, imp, ctr, pos) in enumerate(hourly_data):
    row = 4 + i
    alt = i % 2 == 1
    
    # Add notes for special hours
    note = ""
    if clicks > 0 and ctr > 20:
        note = "⭐ أفضل ساعة - CTR عالي"
    elif clicks > 0:
        note = "✅ نقرات مسجلة"
    elif imp >= 40:
        note = "📈 ذروة الظهور - لكن بدون نقرات!"
    elif imp >= 20:
        note = "📊 ظهور جيد"
    elif pos <= 10 and imp > 0:
        note = "🎯 موضع ممتاز"
    elif imp == 0:
        note = "➖ لا بيانات"
    
    data = [hour, clicks, imp, f"{ctr}%", f"{pos:.1f}", note]
    fill = GRAY_BG if alt else WHITE_BG
    
    for col, val in enumerate(data, 2):
        cell = ws6.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.font = DARK_FONT_SM
        cell.alignment = CENTER
        
        if col == 3 and val > 0:  # Clicks
            cell.font = GREEN_FONT
        if col == 7 and "أفضل" in str(val):
            cell.font = GREEN_FONT
        if col == 7 and "ذروة" in str(val):
            cell.font = RED_FONT

# Hourly summary
h_sum_row = 4 + len(hourly_data) + 2
ws6.merge_cells(f'B{h_sum_row}:G{h_sum_row}')
cell = ws6.cell(row=h_sum_row, column=2, value="📊 ملخص الأداء بالساعة")
cell.font = WHITE_FONT
cell.fill = HEADER_BG
cell.alignment = CENTER
for c in range(2, 8):
    ws6.cell(row=h_sum_row, column=c).fill = HEADER_BG

peak_imp_hour = max(hourly_data, key=lambda x: x[2])
best_ctr_hour = max(hourly_data, key=lambda x: x[3])
best_pos_hour = min((h for h in hourly_data if h[4] > 0), key=lambda x: x[4])

h_summary = [
    ("ساعة ذروة الظهور", f"{peak_imp_hour[0]} ({peak_imp_hour[2]} ظهور)"),
    ("أفضل ساعة CTR", f"{best_ctr_hour[0]} ({best_ctr_hour[3]}%)"),
    ("أفضل موضع", f"{best_pos_hour[0]} (موضع {best_pos_hour[4]})"),
    ("ساعات بنقرات", "16:00 (2 نقرات), 23:00 (1 نقر), 06:00 (1 نقر)"),
    ("ساعات بدون نقرات", "21 من 24 ساعة (87.5%)"),
    ("النمط", "النقرات تحدث في ساعات متفرقة - لا نمط واضح بعد"),
]

for i, (label, val) in enumerate(h_summary):
    r = h_sum_row + 1 + i
    ws6.merge_cells(f'B{r}:D{r}')
    cell = ws6.cell(row=r, column=2, value=label)
    cell.font = DARK_FONT
    cell.fill = LIGHT_BG
    cell.alignment = RTL
    cell.border = THIN_BORDER
    for c in range(2, 5):
        ws6.cell(row=r, column=c).fill = LIGHT_BG
        ws6.cell(row=r, column=c).border = THIN_BORDER
    
    ws6.merge_cells(f'E{r}:G{r}')
    cell = ws6.cell(row=r, column=5, value=val)
    cell.font = Font(color="1B2A4A", size=10)
    cell.fill = WHITE_BG
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    for c in range(5, 8):
        ws6.cell(row=r, column=c).fill = WHITE_BG
        ws6.cell(row=r, column=c).border = THIN_BORDER


# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT_PATH)
print(f"✅ تم إنشاء ملف التحليل بنجاح: {OUTPUT_PATH}")
print(f"📁 الأوراق: {wb.sheetnames}")
